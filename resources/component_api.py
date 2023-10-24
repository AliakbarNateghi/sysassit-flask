import json
from http.client import BAD_REQUEST

import openpyxl
from flask import Blueprint, abort, jsonify, redirect, request, session, url_for
from flask_jwt_extended import get_jwt_identity, jwt_required
from flask_restful import Api, Resource

from ..functions.functions import compare_json
from ..models.models import Component, componentLog, componentUser, db
from ..schemas.schemas import CreateComponentSchema

component_bp = Blueprint("component_api", __name__)
component_api = Api(component_bp)


class component(Resource):
    @jwt_required()
    def put(self, component_id=None):
        user = get_jwt_identity()
        if "username" in session and session["username"] == user:
            user = componentUser.query.filter_by(username=user).first()

        component = Component.query.get(component_id)
        errors = CreateComponentSchema().validate(request.json)
        if errors:
            abort(BAD_REQUEST, str(errors))

        if not component:
            component = Component(id=component_id)

        old_component_data = component.serialize()

        component.section = request.json.get("section")
        component.sub_section = request.json.get("sub_section")
        component.level = request.json.get("level")
        component.TCode = request.json.get("TCode")
        component.code = request.json.get("code")
        component.Parent = request.json.get("Parent")
        component.module = True if request.json.get("module") else False
        component.repeated = True if request.json.get("repeated") else False
        component.condition = request.json.get("condition")
        component.fig = request.json.get("fig")
        component.position = request.json.get("position")
        component.part = request.json.get("part")
        component.quantity = request.json.get("quantity")
        component.standard = request.json.get("standard")
        component.national_stock_number = request.json.get("national_stock_number")
        component.part_number = request.json.get("part_number")
        component.ATA_number = request.json.get("ATA_number")
        component.weight_on_pcs = request.json.get("weight_on_pcs")
        component.weight = request.json.get("weight")
        component.version = request.json.get("version")
        component.sap_code = request.json.get("sap_code")
        component.head_of_CVE = request.json.get("head_of_CVE")
        component.CP = True if request.json.get("CP") else False
        component.CCL = True if request.json.get("CCL") else False
        component.drawing_lvl_2 = True if request.json.get("drawing_lvl_2") else False
        component.ITP_and_SPc_of_material = (
            True if request.json.get("ITP_and_SPc_of_material") else False
        )
        component.ITP_and_SPC_of_production_technology = (
            True if request.json.get("ITP_and_SPC_of_production_technology") else False
        )
        component.type_of_analysis = (
            True if request.json.get("type_of_analysis") else False
        )
        component.prototype_part = True if request.json.get("prototype_part") else False
        component.test = True if request.json.get("test") else False

        new_component_data = component.serialize()

        old_component_data, new_component_data = compare_json(old_component_data, new_component_data)
        
        log = componentLog(
            component_id=component.id,
            user_name=user.username,
            user_email=user.email,
            old_value=old_component_data,
            new_value=new_component_data,
        )

        db.session.add(log)

        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        file_path = "data/GT-22-BOM.xlsx"
        workbook = openpyxl.load_workbook(filename=file_path)
        worksheet = workbook.active

        if component_id:
            component_exel_id = int(component_id) + 1

            worksheet["A" + str(component_exel_id)].value = component_id
            worksheet["B" + str(component_exel_id)].value = component.section
            worksheet["C" + str(component_exel_id)].value = component.sub_section
            worksheet["D" + str(component_exel_id)].value = component.level
            worksheet["E" + str(component_exel_id)].value = component.TCode
            worksheet["F" + str(component_exel_id)].value = component.code
            worksheet["G" + str(component_exel_id)].value = component.Parent
            worksheet["H" + str(component_exel_id)].value = component.module
            worksheet["I" + str(component_exel_id)].value = component.repeated
            worksheet["J" + str(component_exel_id)].value = component.condition
            worksheet["K" + str(component_exel_id)].value = component.fig
            worksheet["L" + str(component_exel_id)].value = component.position
            worksheet["M" + str(component_exel_id)].value = component.part
            worksheet["N" + str(component_exel_id)].value = component.quantity
            worksheet["O" + str(component_exel_id)].value = component.standard
            worksheet[
                "P" + str(component_exel_id)
            ].value = component.national_stock_number
            worksheet["Q" + str(component_exel_id)].value = component.part_number
            worksheet["R" + str(component_exel_id)].value = component.ATA_number
            worksheet["S" + str(component_exel_id)].value = component.weight_on_pcs
            worksheet["T" + str(component_exel_id)].value = component.weight
            worksheet["U" + str(component_exel_id)].value = component.version
            worksheet["V" + str(component_exel_id)].value = component.sap_code
            worksheet["W" + str(component_exel_id)].value = component.head_of_CVE
            worksheet["X" + str(component_exel_id)].value = component.CP
            worksheet["Y" + str(component_exel_id)].value = component.CCL
            worksheet["Z" + str(component_exel_id)].value = component.drawing_lvl_2
            worksheet[
                "AA" + str(component_exel_id)
            ].value = component.ITP_and_SPc_of_material
            worksheet[
                "AB" + str(component_exel_id)
            ].value = component.ITP_and_SPC_of_production_technology
            worksheet["AC" + str(component_exel_id)].value = component.type_of_analysis
            worksheet["AD" + str(component_exel_id)].value = component.prototype_part
            worksheet["AE" + str(component_exel_id)].value = component.test

            file_path = "data/GT-22-BOM.xlsx"
            workbook.save(file_path)
        else:
            # Find the last row of the Excel file and increment its row number by 1
            last_row = worksheet.max_row
            component_exel_id = last_row + 1

            row_values = [
                last_row,
                component.section,
                component.sub_section,
                component.level,
                component.TCode,
                component.code,
                component.Parent,
                component.module,
                component.repeated,
                component.condition,
                component.fig,
                component.position,
                component.part,
                component.quantity,
                component.standard,
                component.national_stock_number,
                component.part_number,
                component.ATA_number,
                component.weight_on_pcs,
                component.weight,
                component.version,
                component.sap_code,
                component.head_of_CVE,
                component.CP,
                component.CCL,
                component.drawing_lvl_2,
                component.ITP_and_SPc_of_material,
                component.ITP_and_SPC_of_production_technology,
                component.type_of_analysis,
                component.prototype_part,
                component.test,
            ]

            worksheet.append(row_values)

            file_path = "data/GT-22-BOM.xlsx"
            workbook.save(file_path)

        # Add the resource to the database and commit changes
        db.session.add(component)
        db.session.commit()

        # Return a JSON response with the updated or created resource
        return component.serialize(), 200

    @jwt_required()
    def get(self, component_id=None):

        component = Component.query.get(component_id)
        if component_id:
            if component:
                component_dict = {
                    c.name: getattr(component, c.name)
                    for c in component.__table__.columns
                }
                json.dumps(component_dict)

                if component_dict:
                    return jsonify(component_dict)
            else:
                return {"message": "component not found"}, 404
        else:
            components = Component.query.all()
            result = []
            for component in components:
                component_dict = {}
                component_dict["id"] = component.id
                component_dict["section"] = component.section
                component_dict["sub_section"] = component.sub_section
                component_dict["level"] = component.level
                component_dict["TCode"] = component.TCode
                component_dict["code"] = component.code
                component_dict["Parent"] = component.Parent
                component_dict["module"] = component.module
                component_dict["repeated"] = component.repeated
                component_dict["condition"] = component.condition
                component_dict["fig"] = component.fig
                component_dict["position"] = component.position
                component_dict["part"] = component.part
                component_dict["quantity"] = component.quantity
                component_dict["standard"] = component.standard
                component_dict[
                    "national_stock_number"
                ] = component.national_stock_number
                component_dict["part_number"] = component.part_number
                component_dict["ATA_number"] = component.ATA_number
                component_dict["weight_on_pcs"] = component.weight_on_pcs
                component_dict["weight"] = component.weight
                component_dict["version"] = component.version
                component_dict["sap_code"] = component.sap_code
                component_dict["head_of_CVE"] = component.head_of_CVE
                component_dict["CP"] = component.CP
                component_dict["CCL"] = component.CCL
                component_dict["drawing_lvl_2"] = component.drawing_lvl_2
                component_dict[
                    "ITP_and_SPc_of_material"
                ] = component.ITP_and_SPc_of_material
                component_dict[
                    "ITP_and_SPC_of_production_technology"
                ] = component.ITP_and_SPC_of_production_technology
                component_dict["type_of_analysis"] = component.type_of_analysis
                component_dict["prototype_part"] = component.prototype_part
                component_dict["test"] = component.test

                result.append(component_dict)

            return jsonify(result)


component_api.add_resource(
    component, "/api/component", "/api/component/<int:component_id>"
)
